# -----------------------------Header information------------------------------
# Author: Ali Reza Shahvaran
# Filename: QA.py
# License: CC BY 4.0
# ----------------------------------------------------------------------------
# Description: This script performs quality assurance analysis on raster data. It includes pixel value counting and data processing for quality assurance purposes.
# ----------------------------------------------------------------------------
# Programming Language: Python 3.0
# Libraries: os, rasterio, pandas, numpy, openpyxl
# ----------------------------------------------------------------------------
# Input: Processes raster data from the directory "C:\\Users\\PHYS3009\\Desktop\\QA\\Inputs".
# ----------------------------------------------------------------------------
# Output: Results are saved to "QA_Results.xlsx" at "C:\\Users\\PHYS3009\\Desktop\\QA\\Outputs".
# ----------------------------------------------------------------------------

import os
import rasterio
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Define directories and file paths for QA1
input_directory = "C:\\Users\\PHYS3009\\Desktop\\QA\\Inputs"
output_excel_path = "C:\\Users\\PHYS3009\\Desktop\\QA\\Outputs\\QA_Results.xlsx"

# Collecting all unique pixel values across all files
all_pixel_values = set()

# Function to count the pixels for QA1
def count_pixels(raster, value):
    return (raster == value).sum()

# Main logic of QA1.py
for File_Name in os.listdir(input_directory):
    if File_Name.endswith(".TIF"):
        file_path = os.path.join(input_directory, File_Name)
        
        # Reading the raster data
        with rasterio.open(file_path) as src:
            raster = src.read(1)
        
        # Adding unique values to the set
        all_pixel_values.update(np.unique(raster))

# Sorting pixel values for column ordering
columns = ['File_Name'] + sorted(list(all_pixel_values))

# Creating a list to hold the results
data = []

# Looping through each file to count pixels for QA1
for File_Name in os.listdir(input_directory):
    if File_Name.endswith(".TIF"):
        file_path = os.path.join(input_directory, File_Name)
        
        # Reading the raster data
        with rasterio.open(file_path) as src:
            raster = src.read(1)
        
        # Counting the pixels for each unique value
        counts = {value: count_pixels(raster, value) for value in all_pixel_values}
        
        # Adding the results to the list
        data.append({'File_Name': File_Name, **counts})

# Creating a DataFrame from the list
df_qa1 = pd.DataFrame(data, columns=columns)

# Saving the DataFrame to an Excel file in a sheet named 'pixel_stats' without the index
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    df_qa1.to_excel(writer, sheet_name='pixel_stats', index=False)

# Functions for QA2.py
def interpret_binary(binary_string):
    flags = {
        "Fill": binary_string[15],
        "Dilated_Cloud": binary_string[14],
        "Cirrus": binary_string[13],
        "Cloud": binary_string[12],
        "Cloud_Shadow": binary_string[11],
        "Snow": binary_string[10],
        "Clear": binary_string[9],
        "Water": binary_string[8],
        "Cloud_Confidence": binary_string[6:8],
        "Cloud_Shadow_Confidence": binary_string[4:6],
        "Snow_Ice_Confidence": binary_string[2:4],
        "Cirrus_Confidence": binary_string[0:2]
    }
    return flags

def header_to_binary(header):
    return format(int(header), '016b')

# Main logic of QA2.py
wb = load_workbook(output_excel_path)
sheet = wb.active

# Get headers (including the first numeric one)
headers = [cell.value for cell in sheet[1]][1:]

# Initialize interpreted_data as an empty list
interpreted_data = []

# Main logic of QA2.py
for header in headers:  # Start processing from the first numeric header
    binary_header = header_to_binary(header)
    flags = interpret_binary(binary_header)
    interpreted_data.append({"Original_Header": header, "Binary_Format": binary_header, **flags})


# Creating a DataFrame from the interpreted data
df_qa2 = pd.DataFrame(interpreted_data)

# Saving the DataFrame of QA2.py to the same Excel file in a sheet named 'pixel_descr' without the index
with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a') as writer:
    df_qa2.to_excel(writer, sheet_name='pixel_descr', index=False)

# Step 1: Duplicate the pixel_stats DataFrame
df_qa1_percentage = df_qa1.copy()

# Step 2: Remove the first two columns ('File_Name' and '1')
df_qa1_percentage.drop(columns=['File_Name', 1], inplace=True)

# Step 3: Calculate the percentage of each cell relative to the sum of its row
df_qa1_percentage = df_qa1_percentage.div(df_qa1_percentage.sum(axis=1), axis=0) * 100

# Step 4: Reinsert the 'File_Name' column to the beginning of the DataFrame
df_qa1_percentage.insert(0, 'File_Name', df_qa1['File_Name'])
df_qa1_percentage = df_qa1_percentage.round(1)
# Saving the DataFrame with percentage values to the Excel file
with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a') as writer:
    df_qa1_percentage.to_excel(writer, sheet_name='pixel_stats2', index=False)
    
# Definitions for the new descriptions
descriptions = {
    "Fill": ["Image data", "Fill data"],
    "Dilated_Cloud": ["Cloud is not dilated or no cloud", "Cloud dilation"],
    "Cirrus": ["Cirrus confidence: no confidence level set or low confidence", "High confidence cirrus"],
    "Cloud": ["Cloud confidence is not high", "High confidence cloud"],
    "Cloud_Shadow": ["Cloud shadow confidence is not high", "High confidence cloud shadow"],
    "Snow": ["Snow/Ice confidence is not high", "High confidence snow cover"],
    "Clear": ["Cloud or dilated cloud bits are set", "Cloud and dilated cloud bits are not set"],
    "Water": ["Land or cloud", "Water"],
    "Cloud_Confidence": ["No confidence level set", "Low confidence", "Medium confidence", "High confidence"],
    "Cloud_Shadow_Confidence": ["No confidence level set", "Low confidence", "Reserved", "High confidence"],
    "Snow_Ice_Confidence": ["No confidence level set", "Low confidence", "Reserved", "High confidence"],
    "Cirrus_Confidence": ["No confidence level set", "Low confidence", "Reserved", "High confidence"]
}

# Duplicating the pixel_descr DataFrame
df_qa2_descr = df_qa2.copy()

# Modifying the values according to the descriptions
for column in descriptions.keys():
    if column in ["Cloud_Confidence", "Cloud_Shadow_Confidence", "Snow_Ice_Confidence", "Cirrus_Confidence"]:
        # Special handling for two-digit binary columns
        mapping_dict = {
            '00': descriptions[column][0],
            '01': descriptions[column][1],
            '10': descriptions[column][2],
            '11': descriptions[column][3]
        }
        df_qa2_descr[column] = df_qa2[column].astype(str).map(mapping_dict)
    else:
        # Mapping single binary values to descriptions for other columns
        df_qa2_descr[column] = df_qa2[column].map(lambda x: descriptions[column][int(x)])

# Saving the modified DataFrame to the same Excel file in a new sheet named 'pixel_descr2'
with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a') as writer:
    df_qa2_descr.to_excel(writer, sheet_name='pixel_descr2', index=False)

from openpyxl.styles import Alignment

# Function to auto-adjust column widths and set alignment
def adjust_column_widths_and_align(worksheet):
    for column_cells in worksheet.columns:
        length_max = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length_max + 2
        
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Saving the DataFrames to the Excel file and adjusting column widths and alignment
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    # Saving QA1 DataFrame
    df_qa1.to_excel(writer, sheet_name='pixel_stats', index=False)
    
    df_qa1_percentage.to_excel(writer, sheet_name='pixel_stats2', index=False)
    
    # Saving QA2 DataFrame
    df_qa2.to_excel(writer, sheet_name='pixel_descr', index=False)
    
    # Saving the modified QA2 DataFrame
    df_qa2_descr.to_excel(writer, sheet_name='pixel_descr2', index=False)
    
    # Getting the workbook object
    workbook = writer.book
    
    # Adjusting column widths and alignment for each sheet
    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        adjust_column_widths_and_align(worksheet)


